"""
These utils generate metasummaries and illustrative quotes
from the batch-checked quotes and answers to research questions.
"""

import os

from typing import Dict
from typing import List
from typing import Tuple
from typing import Optional

import pandas as pd

from langchain.chat_models import AzureChatOpenAI
from langchain.prompts import PromptTemplate
from langchain.output_parsers import PydanticOutputParser
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from pydantic import BaseModel, Field


llm = AzureChatOpenAI(
    openai_api_version=os.getenv("AZURE_OPENAI_API_VERSION"),
    azure_deployment=os.getenv("AZURE_OPENAI_DEPLOYMENT_NAME"),
    azure_endpoint=os.getenv("AZURE_OPENAI_ENDPOINT"),
    openai_api_key=os.getenv("AZURE_OPENAI_API_KEY"),
    temperature=0,
)


class QuoteSelectionOutput(BaseModel):
    """Pydantic model for structured output representing selected quotes."""

    quotes: List[str] = Field(
        description="A list of exactly 5 quotes, copied verbatim from the input. Each quote must be an exact match, with no modifications."
    )


quote_parser = PydanticOutputParser(pydantic_object=QuoteSelectionOutput)

SUMMARY_PROMPT = PromptTemplate.from_template(
    """
    You are assisting with qualitative research by summarizing multiple short answers to a research question.

    Each answer below is derived from a single document and is supported by direct, verbatim quotes from that document.

    Research question: '{question}'

    Based on the following answers, generate a concise, high-level summary that answers the research question.
    Only use the information contained in the input answers. Do not introduce new ideas, assumptions, or background knowledge.

    Answers:
    {context}
    """
)

QUOTE_PROMPT = PromptTemplate.from_template(
    "Your job is to return quotes without modifying them in any way."
    "Here are some verbatim quotes from transcripts:\n\n{context}\n\n"
    "The answer to the question '{question}' was: {answer}\n\n"
    "Select exactly 5 quotes that best support this answer.\n"
    "Quotes must be copied verbatim. Do not change them in any way."
    "Return output in this format: \n {format_instructions}"
)


def summarize_and_quote(texts: List[str], answers: List[str], question: str) -> Tuple[str, List[str]]:
    """
    Given a set of answers to a research question that were each based on one document,
    synthesise these into a metasummary.

    Then select 5 quotes from the texts (quotes that were also selected at the same stage that these
    single-document answers were generated) that best support the metasummary.

    Args:
        texts (List[str]): List of quotes from batch_check output
        answers (List[str]): List of answers from batch_check output
        question (str): Research question to be answered

    Returns:
        Tuple[str, List[str]]: Metasummary answer to the question, and supporting best quotes
    """

    # Step 1: Generate the summary
    answer_response = llm.invoke(SUMMARY_PROMPT.format(context="\n\n".join(answers), question=question))
    answer = answer_response.content.strip()

    # Step 2: Select supporting quotes
    quote_prompt_filled = QUOTE_PROMPT.partial(format_instructions=quote_parser.get_format_instructions())
    quote_input = {
        "context": "\n\n".join(texts),
        "question": question,
        "answer": answer,
    }
    formatted_prompt = quote_prompt_filled.format(**quote_input)

    try:
        quote_response = llm.invoke(formatted_prompt)
        parsed_output: QuoteSelectionOutput = quote_parser.parse(quote_response.content)
        quotes_list = parsed_output.quotes
    except Exception as e:
        print(f"⚠️ Quote parsing failed: {e}")
        quotes_list = ["[Parsing error]"]

    return answer, quotes_list


def generate_summaries_and_quotes(
    rq_dict: Dict[str, str], batch_check_all_rqs: pd.DataFrame
) -> Dict[str, Dict[str, object]]:
    """
    Generate summary answers and supporting quotes for each research question.

    For each research question in `rq_dict`, this function:
    - Filters the corresponding rows from the batch-checked quotes DataFrame,
    - Extracts the list of supporting texts and answers,
    - Passes them to `summarize_and_quote` to get a summary and selected quotes.

    Args:
        rq_dict (Dict[str, str]):
            Dictionary mapping RQ IDs (e.g., "rq_1") to research question strings.
        batch_check_all_rqs (pd.DataFrame):
            DataFrame containing validated quotes with columns 'rq', 'answer', and 'text'.

    Returns:
        Dict[str, Dict[str, object]]:
            A nested dictionary mapping each RQ ID to a dict with keys:
            - 'answer': the generated summary,
            - 'quotes': a list of supporting quotes.
    """
    summarize_and_quote_output = {}

    for id, question in rq_dict.items():
        batch_check_output = batch_check_all_rqs[batch_check_all_rqs["rq"] == question]

        answers = batch_check_output["answer"].tolist()

        extracted_texts = batch_check_output["text"].tolist()

        answer, quotes_list = summarize_and_quote(extracted_texts, answers, question)

        summarize_and_quote_output[id] = {"answer": answer, "quotes": quotes_list}

    return summarize_and_quote_output


def identify_source_of_quote(
    conversation_text_dict: Dict[str, str], quotes_list: List[str]
) -> Dict[str, Optional[str]]:
    """
    Generate a mapping of each quote returned by `generate_summaries_and_quotes`
    to the filename/conversation id it was found in.

    For each quote, this function searches all values in the conversation_text_dict
    to find which conversation (file) it came from.

    Args:
        conversation_text_dict (Dict[str, str]):
            Dictionary mapping file/conversation IDs to full transcript text.
        quotes_list (List[str]):
            List of quotes to locate.

    Returns:
        Dict[str, Optional[str]]:
            A dictionary mapping each quote to the filename it was found in,
            or `None` if not found.
    """
    quote_to_file = {}

    for quote in quotes_list:
        found = False
        for filename, full_text in conversation_text_dict.items():
            if quote in full_text:
                quote_to_file[quote] = filename
                found = True
                break
        if not found:
            quote_to_file[quote] = None

    for quote, file in quote_to_file.items():
        print(f"QUOTE:\n{quote}\n→ FOUND IN: {file}\n")

    return quote_to_file


def create_final_output_df(
    rq_dict: Dict[str, str],
    conversation_text_dict: Dict[str, str],
    summarize_and_quote_output: Dict[str, Dict[str, object]],
) -> pd.DataFrame:
    """
    Compile final output DataFrame containing research questions, summary answers, top quotes,
    and their sources (file names or conversation ids).

    For each RQ:
    - Matches selected quotes to their original source file,
    - *Removes* quotes that could not be matched to a source,
                (i.e. quotes that may have been hallucinated)
    - Adds metadata such as the RQ ID, full question, and summary answer.

    Args:
        rq_dict (Dict[str, str]):
            Dictionary mapping RQ IDs to research question strings.
        conversation_text_dict (Dict[str, str]):
            Dictionary mapping file/conversation IDs to full transcript text.
        summarize_and_quote_output (Dict[str, Dict[str, object]]):
            Output from `generate_summaries_and_quotes`, containing:
              - 'answer': generated summary,
              - 'quotes': selected supporting quotes.

    Returns:
        pd.DataFrame:
            A DataFrame with columns ['question', 'answer', 'quote', 'source'].
    """
    summarise_output_df = pd.DataFrame()

    for rq_id, question in rq_dict.items():
        print(f"RQ ID: {rq_id}")

        quote_to_file_mapping = identify_source_of_quote(
            conversation_text_dict, summarize_and_quote_output[rq_id]["quotes"]
        )
        quote_mapping_df = pd.DataFrame(list(quote_to_file_mapping.items()), columns=["quote", "source"])
        quote_mapping_df = quote_mapping_df[quote_mapping_df["source"].notnull()]
        quote_mapping_df["rq_id"] = rq_id
        quote_mapping_df["question"] = question
        quote_mapping_df["answer"] = summarize_and_quote_output[rq_id]["answer"]
        summarise_output_df = pd.concat([summarise_output_df, quote_mapping_df], ignore_index=True)

    return summarise_output_df[["question", "answer", "quote", "source"]]


def merge_column(ws: Worksheet, col_idx: int, start_row: int = 2) -> None:
    """
    Merge rows within a column that have the same value.

    The start row is 2 because row 1 is assumed to be the header.
    """
    row = start_row
    while row <= ws.max_row:
        current_value = ws.cell(row=row, column=col_idx).value
        end_row = row
        while end_row + 1 <= ws.max_row and ws.cell(row=end_row + 1, column=col_idx).value == current_value:
            end_row += 1
        if end_row > row:
            ws.merge_cells(start_row=row, start_column=col_idx, end_row=end_row, end_column=col_idx)
            ws.cell(row=row, column=col_idx).alignment = Alignment(vertical="center")
        row = end_row + 1


def create_output_excel(full_summary_df: pd.DataFrame, output_dir: str) -> None:
    """
    Save the full output (answers to RQs and illustrative quotes) to excel.
    This excel file is then available for download in the Dash session.

    Args:
        full_summary_df (pd.DataFrame): The full summary output DataFrame.
        output_dir (str): Directory to save the Excel file to.
    """
    df = full_summary_df.copy()

    # List of columns to merge
    columns_to_merge = ["question", "answer"]

    # Save DataFrame to Excel
    output_path = f"{output_dir}/full_summary.xlsx"
    df.to_excel(output_path, index=False)

    # Load with openpyxl
    wb = load_workbook(output_path)
    ws = wb.active

    # Apply merge to each column in `columns_to_merge`
    for col_name in columns_to_merge:
        col_idx = list(df.columns).index(col_name) + 1  # Convert to Excel's 1-indexing
        merge_column(ws, col_idx)

    # Save the result
    wb.save(output_path)
